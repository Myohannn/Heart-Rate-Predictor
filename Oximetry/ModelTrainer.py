import win32ui
import xlwt
import pandas
import pickle
import pandas as pd

import numpy as np
import openpyxl

from keras.callbacks import EarlyStopping, ModelCheckpoint

from tensorflow.keras.layers import Dense
from tensorflow.keras.models import Sequential

from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split

from sklearn.metrics import mean_squared_error

from sklearn.ensemble import RandomForestClassifier

import configparser

conf = configparser.ConfigParser()
conf.read('config.ini')


class ModelTrainer:
    def __init__(self, mode):
        if mode == "1":
            filename = self.get_file()
            RFClassifierTrainer(filename).run()
        elif mode == "2":
            filename = self.get_file()
            RFPredictorTrainer(filename).run()
        elif mode == "3":
            filename = self.get_file()
            KerasNNPredictorTrainer(filename).run()
        else:
            print("Invalid input")

    def get_file(self):
        def read_file():
            dlg = win32ui.CreateFileDialog(1)
            # dlg.SetOFNInitialDir(r'C:\Users\csdwhuang\Desktop\PPG_model\Oximetry\dataset\training_data')
            dlg.DoModal()
            filename = dlg.GetPathName()
            return filename

        print("Please select training file(.xlsx)")
        train_file_path = read_file()
        print("Selected：{}\n".format(train_file_path))
        return train_file_path


class RFClassifierTrainer:
    def __init__(self, filename):
        self.filename = filename

    def load_data(self):
        dataset = pd.read_excel(self.filename)
        X = dataset.iloc[:, :4]
        y = dataset.iloc[:, 4].values
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=conf.get("Model_Trainer",
                                                                                     "clf_test_size"), random_state=0,
                                                            shuffle=True)
        return X_train, X_test, y_train, y_test, X

    def biuld_model(self, datas):
        clf = RandomForestClassifier(n_estimators=conf.get("Model_Trainer", "clf_n_estimators"))
        clf.fit(datas[0], datas[2])
        print("Accuracy(RF)：%.4lf" % clf.score(datas[1], datas[3]))
        # save the model to disk
        pickle.dump(clf, open("model/classifier/HR_Level_Classifier.pkl", 'wb'))
        return clf

    def prediction(self, classifier, para):
        target = classifier.predict(np.array(para))
        return float(target)

    def auto_prediction(self, classifier, dataset):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet(index=0)

        worksheet.cell(1, 5).value = 'Action'
        data = np.array(dataset)

        i = 1
        for n in data:
            target = self.prediction(classifier, np.array([n]))
            worksheet.cell(i, 1).value = float(n[0])
            worksheet.cell(i, 2).value = float(n[1])
            worksheet.cell(i, 3).value = float(n[2])
            worksheet.cell(i, 4).value = float(n[3])
            worksheet.cell(i, 5).value = float(target)
            i += 1
            # print(i)
        workbook.save('dataset/training_result/randomForest_140.csv')

    def run(self):
        print("Loading dataset......\n")
        X_train, X_test, y_train, y_test, X = self.load_data()
        datas = [X_train, X_test, y_train, y_test, X]
        print("Biulding model......\n")
        classifier = self.biuld_model(datas)

        # print("Evaluating model......\n")
        # self.auto_prediction(classifier, X)


class RFPredictorTrainer:
    def __init__(self, filename):
        self.filename = filename

    def run(self):
        data = pd.read_excel(self.filename)
        print("Loading dataset......\n")
        X = data[[' Green Count', ' Green2 Count', ' IR Count', ' Red Count']]
        Y = data[['record.heart_rate[bpm]']]

        X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=conf.get("Model_Trainer",
                                                                                     "RFpredictor_test_size"), random_state=42, shuffle=True)

        print("Biulding model......\n")
        rf_model = RandomForestRegressor(n_estimators=conf.get("Model_Trainer", "RFpredictor_n_estimators"))
        rf_model.fit(X_train, y_train)
        pickle.dump(rf_model, open("model/predictor/RandomForest/RF_Predictor_model.pkl", 'wb'))
        print("Model(RF_Predictor_model.pkl) saved......\n")
        y_predict = rf_model.predict(X_test)
        mse = mean_squared_error(y_predict, y_test)
        print("MSE of RF Predictor is ", mse)


class KerasNNPredictorTrainer:
    def __init__(self, filename):
        data = pandas.read_excel(filename)

        self.X = data[[' Green Count', ' Green2 Count', ' IR Count', ' Red Count']]
        self.Y = data[['record.heart_rate[bpm]']]

        self.X_train, self.X_test, self.y_train, self.y_test = train_test_split(self.X, self.Y,
                                                                                test_size=conf.get("Model_Trainer",
                                                                                                   "keras_test_size"),
                                                                                random_state=42,
                                                                                shuffle=True)

        self.result_dict = {}

    def biuld_model(self, units, layers):
        # Initialising the ANN
        model = Sequential()

        # Adding the input layer and the first hidden layer
        model.add(Dense(units=units, input_dim=4, activation='relu'))  # second layer overall

        for i in range(0, layers - 2):
            # Adding hidden layer
            model.add(Dense(units=units, activation='relu'))

        # Adding the output layer
        model.add(Dense(units=1))
        # model.summary()  # log

        # model.compile(optimizer='adam', loss='mean_squared_error')
        model.compile(optimizer='rmsprop', loss='mean_squared_error')
        return model

    def train_model(self, units, layers, batch, epoch):
        model = self.biuld_model(units, layers)
        print(
            "Units:" + str(units) + " Layers: " + str(layers) + " Epoch: " + str(epoch) + " Batch: " + str(
                batch) + " is training......")

        early_stopping = EarlyStopping(monitor='val_loss', patience=50, mode='min', verbose=0)

        result_key = '' + str(units) + '_' + str(layers) + '_' + str(batch) + '_' + str(epoch)

        model_name = 'model/predictor/KerasNN/keras_model_' + result_key + ".h5"

        mc = ModelCheckpoint(filepath=model_name,
                             monitor='val_loss',
                             mode='min',
                             verbose=2,
                             save_best_only=True)

        history = model.fit(self.X_train, self.y_train, batch_size=batch, epochs=epoch, verbose=1,
                            validation_split=0.2, callbacks=[mc, early_stopping])

        loss = model.evaluate(self.X_test, self.y_test, verbose=0)

        self.result_dict[result_key] = loss
        print("Test loss(MSE):", loss)
        # print()

    def post_result(self, result):
        sorted_result = sorted(result.items(), key=lambda kv: (kv[1], kv[0]))
        print(sorted_result)

        workbook = xlwt.Workbook(encoding='utf-8')
        worksheet = workbook.add_sheet('result')

        worksheet.write(0, 0, label='Parameter')
        worksheet.write(0, 1, label='Unit')
        worksheet.write(0, 2, label='Layer')
        worksheet.write(0, 3, label='Batch_size')
        worksheet.write(0, 4, label='Epoch')
        worksheet.write(0, 5, label='MSE')

        for i in range(0, len(sorted_result)):
            parameters = sorted_result[i][0].split("_")
            worksheet.write(i + 1, 0, label=sorted_result[i][0])
            worksheet.write(i + 1, 1, label=parameters[0])
            worksheet.write(i + 1, 2, label=parameters[1])
            worksheet.write(i + 1, 3, label=parameters[2])
            worksheet.write(i + 1, 4, label=parameters[3])
            worksheet.write(i + 1, 5, label=sorted_result[i][1])

        workbook.save('dataset/training_result/model_training_result.xls')

    def run(self):

        # config file
        units_Array = conf.get("Model_Trainer", "units_Array")
        layers_Array = conf.get("Model_Trainer", "layers_Array")
        batch_Array = conf.get("Model_Trainer", "batch_Array")
        epoch_Array = conf.get("Model_Trainer", "epoch_Array")

        for units in units_Array:
            for layers in layers_Array:
                for batch in batch_Array:
                    for epoch in epoch_Array:
                        self.train_model(units, layers, batch, epoch)

        print()
        print("Saving training result......")
        self.post_result(self.result_dict)
        print("Training result saved!")
        # trainset=[(64,8,4,400),(32,8,8,500),(64,8,12,200),(64,8,12,300),(64,8,6,500),(128,6,12,300),(32,12,6,200),(64,8,12,500)]
        # trainset = [(64, 8, 12, 10),(64, 8, 12, 7),(64, 8, 12, 5)]
        # for set in trainset:
        #     train_model(set[0], set[1], set[2], set[3])

        # print()
        # print(self.result_dict)
