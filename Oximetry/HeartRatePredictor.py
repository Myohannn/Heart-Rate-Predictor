import pickle
import numpy as np
import openpyxl
import pandas as pd
import win32ui

# Read the dataset
from tensorflow.python.keras.models import load_model

import configparser

conf = configparser.ConfigParser()
conf.read('config.ini')


class HRpredictor:
    def __init__(self):
        self.filename = self.get_file("dataset:")
        self.clfname = self.get_file("heart rate level classifier:")
        self.model_low = self.get_file("model for low level heart rate:")
        self.model_medium = self.get_file("model for medium level heart rate:")
        self.model_high = self.get_file("model for high level heart rate:")
        self.models = []

    def get_file(self, hint):

        def read_file():
            dlg = win32ui.CreateFileDialog(1)
            # dlg.SetOFNInitialDir(r'C:\Users\csdwhuang\Desktop\PPG_model\Oximetry\dataset\training_data')
            dlg.DoModal()
            filename = dlg.GetPathName()
            return filename

        print("Please select", hint)
        train_file_path = read_file()
        print("Selected：{}\n".format(train_file_path))
        return train_file_path

    def load_dataset(self):
        dataset = pd.read_excel(self.filename)
        X = dataset[[' Green Count', ' Green2 Count', ' IR Count', ' Red Count']]
        # X = dataset.iloc[:, :4]
        # y = dataset.iloc[:, 4].values
        # data = np.array(X)
        return X

    def predict_class(self, para):
        # clf = load_model(self.clfname)
        clf = pickle.load(open(self.clfname, 'rb'))

        target = clf.predict(np.array([para]))
        return int(target)

    def load_prediction_model(self):
        # load models
        if 'h5' in self.model_low:
            model1 = load_model(self.model_low)
        elif 'pkl' in self.model_low:
            model1 = pickle.load(open(self.model_low, 'rb'))
        else:
            model1 = "Invalid low model"

        if 'h5' in self.model_medium:
            model2 = load_model(self.model_medium)
        elif 'pkl' in self.model_medium:
            model2 = pickle.load(open(self.model_medium, 'rb'))
        else:
            model2 = "Invalid medium model"

        if 'h5' in self.model_high:
            model3 = load_model(self.model_high)
        elif 'pkl' in self.model_high:
            model3 = pickle.load(open(self.model_high, 'rb'))
        else:
            model3 = "Invalid high model"
        self.models = [model1, model2, model3]

        return (isinstance(model1, str) or isinstance(model2, str) or isinstance(model3, str))

    def predict_HR(self, data):
        try:
            level = self.predict_class(data)
            unknown = np.array([data], dtype=np.float32)
            HR = self.models[level].predict(unknown)
        except Exception as e:
            print(e)
            HR = level = "Classification Error"
            print("Classification Error")

        return level, HR

    def run_dataset(self):
        # load models
        if not self.load_prediction_model():
            # load dataset
            data = self.load_dataset()
            workbook = openpyxl.Workbook()
            worksheet = workbook.create_sheet(index=0)
            worksheet.cell(1, 1).value = 'Green Count'
            worksheet.cell(1, 2).value = 'Green2 Count'
            worksheet.cell(1, 3).value = 'IR Count'
            worksheet.cell(1, 4).value = 'Red Count'
            worksheet.cell(1, 5).value = 'HR_level'
            worksheet.cell(1, 6).value = 'HR'
            data = np.array(data)

            i = 2
            HR_low = int(conf.get("HR_Predictor", "HR_low"))
            HR_high = int(conf.get("HR_Predictor", "HR_high"))
            HR_level = [HR_low + "-", HR_low + "~" + HR_high, HR_high + "+"]

            print("Predicting heart rate......")
            for n in data:
                validate_bool = False
                level, HR = self.predict_HR(n)
                if 50 < HR <= int(HR_low):
                    validate_level = 0
                elif int(HR_low) < HR < int(HR_high):
                    validate_level = 1
                elif int(HR_high) <= HR < 190:
                    validate_level = 2
                else:
                    validate_level = -1

                if level == validate_level:
                    validate_bool = True

                worksheet.cell(i, 1).value = float(n[0])
                worksheet.cell(i, 2).value = float(n[1])
                worksheet.cell(i, 3).value = float(n[2])
                worksheet.cell(i, 4).value = float(n[3])
                worksheet.cell(i, 5).value = HR_level[int(level)]

                index = str(i - 2) + ":"
                if validate_bool:
                    print(index, "The predicted heart rate is", HR)
                    worksheet.cell(i, 6).value = float(HR)
                else:
                    print(index, "The predicted heart rate is bad")
                    worksheet.cell(i, 6).value = "bad_" + str(HR)

                i += 1

            print("Predicting result saved as: HR_prediction.xlsx")
            workbook.save('HR_prediction.xlsx')
        else:
            print("Invalid selected model")

    # def run_data(self, data):
    #     # load classifier
    #     clf = pickle.load(open(self.clfname, 'rb'))
    #     # load dataset
    #     for i in data:
    #         level, HR = self.predict_HR(i)
    #         HR_level = ["Level1:100-", "Level2:100~140", "Level3:140+"]
    #         print("It belongs to ", HR_level[int(level)])
    #         print("\nThe predicted heart rate is", HR)

    # def run(self):
    #     print("Loading models......\n")
    #     # datas = [X_train, X_test, y_train, y_test, X]
    #     print("Biulding model......\n")
    #     classifier = self.biuld_model(datas)

    # print("Evaluating model......\n")
    # self.auto_prediction(classifier, X)
